import importlib
import os
import sys
import tempfile
import unittest
from datetime import date, datetime, time, timedelta
from io import BytesIO

from openpyxl import load_workbook


class EngineeringFlowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        cls.db_path = os.path.join(cls.temp_dir.name, "engineering_test.db")
        os.environ["DATABASE_URL"] = f"sqlite:///{cls.db_path}"
        os.environ["SECRET_KEY"] = "engineering-test-secret"
        os.environ["APP_TIMEZONE"] = "Asia/Kolkata"
        os.environ["ENABLE_BACKGROUND_JOBS"] = "0"
        os.environ["APP_STORAGE_ROOT"] = os.path.join(cls.temp_dir.name, "uploads")
        os.environ["APP_BACKUP_ROOT"] = os.path.join(cls.temp_dir.name, "backups")

        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        if project_dir not in sys.path:
            sys.path.insert(0, project_dir)

        if "app" in sys.modules:
            cls.app_module = importlib.reload(sys.modules["app"])
        else:
            cls.app_module = importlib.import_module("app")

        cls.app = cls.app_module.app
        cls.db = cls.app_module.db
        cls.app.config["TESTING"] = True

    @classmethod
    def tearDownClass(cls):
        cls.temp_dir.cleanup()

    def setUp(self):
        self.client = self.app.test_client()
        with self.app.app_context():
            self.db.drop_all()
            self.db.create_all()
            self._seed_admin()
        fallback_file = getattr(self.app_module, "HOLD_BILL_FALLBACK_FILE", "")
        if fallback_file and os.path.exists(fallback_file):
            os.remove(fallback_file)

    def _seed_admin(self):
        admin = self.app_module.User(
            username="admin",
            role="admin",
            access_profile="admin",
            can_manage_users=True,
            can_view_reports=True,
            can_invoice_action=True,
            can_edit_invoice=True,
            can_delete_invoice=True,
            can_view_medicine=True,
            can_add_medicine=True,
            can_edit_medicine=True,
            can_delete_medicine=True,
            can_view_stock_history=True,
            can_manage_purchases=True,
            can_view_audit_logs=True,
            can_view_profit_dashboard=True,
        )
        admin.set_password("Admin@123")
        self.db.session.add(admin)
        self.db.session.commit()

    def login(self):
        return self.login_as("admin", "Admin@123")

    def login_as(self, username, password):
        response = self.client.post(
            "/login",
            data={"username": username, "password": password},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        return response

    def _seed_patient(self, name="Ravi Kumar", mobile="9876543210"):
        patient = self.app_module.Patient(name=name, mobile=mobile, gender="MALE", age=35)
        self.db.session.add(patient)
        self.db.session.flush()
        return patient

    def _seed_vendor_purchase_stack(self, *, medicine_name="PARACETAMOL 650", batch="B123", total_qty=10, purchase_chunks=None):
        if purchase_chunks is None:
            purchase_chunks = [(total_qty, 10.0, datetime.utcnow())]
        vendor = self.app_module.Vendor(name="Prime Distributor")
        self.db.session.add(vendor)
        self.db.session.flush()

        medicine = self.app_module.Medicine(
            name=medicine_name,
            batch=batch,
            expiry="2027-12-31",
            mrp=25.0,
            qty=total_qty,
            discount_percent=5,
            barcode="PARA650",
            reorder_level=10,
            is_active=True,
        )
        self.db.session.add(medicine)
        self.db.session.flush()

        purchase = self.app_module.VendorPurchase(
            vendor_id=vendor.id,
            purchase_no="PB-000001",
            invoice_no="SUP-1001",
            purchase_date=datetime.utcnow(),
            payment_mode="CASH",
            payment_status="Paid",
            total_amount=sum(qty * rate for qty, rate, _ in purchase_chunks),
            created_by="admin",
        )
        self.db.session.add(purchase)
        self.db.session.flush()

        for qty, rate, created_at in purchase_chunks:
            self.db.session.add(
                self.app_module.VendorPurchaseItem(
                    purchase_id=purchase.id,
                    vendor_id=vendor.id,
                    medicine_id=medicine.id,
                    medicine_name=medicine.name,
                    barcode=medicine.barcode,
                    batch=batch,
                    expiry=medicine.expiry,
                    qty=qty,
                    free_qty=0,
                    remaining_qty=qty,
                    purchase_rate=rate,
                    mrp=medicine.mrp,
                    total_value=qty * rate,
                    created_at=created_at,
                )
            )
        self.db.session.commit()
        return vendor, purchase, medicine

    def _seed_invoice(
        self,
        *,
        customer="Ravi Kumar",
        customer_gst_no="",
        mobile="9876543210",
        total=50.0,
        subtotal=None,
        payment_mode="CASH",
        cash_amount=None,
        online_amount=None,
        is_split_payment=False,
        internal_note="",
        created_by="admin",
        created_at=None,
    ):
        patient = self._seed_patient(name=customer, mobile=mobile)
        invoice_count = self.app_module.Invoice.query.count() + 1
        payable_total = float(total if total is not None else 0)
        rounded_total = self.app_module.compute_invoice_rounded_total(payable_total)
        if subtotal is None:
            subtotal = payable_total
        normalized_payment_mode = (payment_mode or "").strip().upper()
        if cash_amount is None:
            cash_amount = rounded_total if normalized_payment_mode == "CASH" else 0.0
        if online_amount is None:
            if normalized_payment_mode in {"CASH", "ADJUSTMENT", "CREDIT"}:
                online_amount = 0.0
            else:
                online_amount = rounded_total
        if created_at is None:
            created_at = datetime.utcnow()

        invoice = self.app_module.Invoice(
            invoice_no=f"INV-SEED-{invoice_count}",
            patient_id=patient.id,
            customer=customer,
            customer_gst_no=customer_gst_no,
            mobile=mobile,
            subtotal=subtotal,
            total=payable_total,
            payment_mode=payment_mode,
            cash_amount=cash_amount,
            online_amount=online_amount,
            is_split_payment=is_split_payment,
            internal_note=internal_note,
            created_by=created_by,
            created_at=created_at,
        )
        self.db.session.add(invoice)
        self.db.session.commit()
        return patient, invoice

    def _expected_auto_code(self, name, serial=1):
        return f"{self.app_module.build_medicine_code_prefix(name)}{serial:03d}"

    def test_billing_flow_creates_invoice_and_updates_stock(self):
        with self.app.app_context():
            self._seed_patient()
            self._seed_vendor_purchase_stack(total_qty=10)

        self.login()
        response = self.client.post(
            "/billing",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "CASH",
                "internal_note": "Handle quietly at counter",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["2"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"INV-", response.data)

        with self.app.app_context():
            invoice = self.app_module.Invoice.query.one()
            invoice_item = self.app_module.InvoiceItem.query.one()
            medicine = self.app_module.Medicine.query.filter_by(name="PARACETAMOL 650", batch="B123").one()
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            stock_history = self.app_module.StockHistory.query.filter_by(action="SALE").all()

            self.assertEqual(invoice.customer, "Ravi Kumar")
            self.assertEqual(float(invoice.cash_amount), 48.0)
            self.assertEqual(float(invoice.online_amount), 0.0)
            self.assertFalse(invoice.is_split_payment)
            self.assertEqual(invoice.internal_note, "Handle quietly at counter")
            self.assertEqual(invoice_item.qty, 2)
            self.assertEqual(medicine.qty, 8)
            self.assertEqual(purchase_item.remaining_qty, 8)
            self.assertEqual(len(stock_history), 1)
        self.assertNotIn(b"Handle quietly at counter", response.data)

    def test_billing_flow_records_split_payment_breakdown(self):
        with self.app.app_context():
            self._seed_patient()
            self._seed_vendor_purchase_stack(total_qty=10)

        self.login()
        response = self.client.post(
            "/billing",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "UPI",
                "split_cash_amount": "18.00",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["2"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Cash Paid", response.data)
        self.assertIn(b"Online Paid", response.data)

        with self.app.app_context():
            invoice = self.app_module.Invoice.query.one()
            self.assertEqual(invoice.payment_mode, "UPI")
            self.assertEqual(float(invoice.cash_amount), 18.0)
            self.assertEqual(float(invoice.online_amount), 30.0)
            self.assertTrue(invoice.is_split_payment)

    def test_billing_flow_starts_fresh_start_invoice_series_from_one(self):
        with self.app.app_context():
            self._seed_patient()
            self._seed_vendor_purchase_stack(total_qty=12)
            legacy_style_invoice = self.app_module.Invoice(
                invoice_no="INV-2026-1622",
                customer="Already Created",
                mobile="9000000100",
                subtotal=50.0,
                total=50.0,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime(2026, 7, 2, 10, 0, 0),
            )
            self.db.session.add(legacy_style_invoice)
            self.db.session.commit()

        self.login()
        first_response = self.client.post(
            "/billing",
            data={
                "customer": "Fresh Start One",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "CASH",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["1"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(first_response.status_code, 200)
        self.assertIn(b"INV-00001", first_response.data)

        second_response = self.client.post(
            "/billing",
            data={
                "customer": "Fresh Start Two",
                "mobile": "9876543211",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "CASH",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["1"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(second_response.status_code, 200)
        self.assertIn(b"INV-00002", second_response.data)

        with self.app.app_context():
            invoice_nos = [
                row[0]
                for row in self.db.session.query(self.app_module.Invoice.invoice_no)
                .order_by(self.app_module.Invoice.id.asc())
                .all()
            ]
            self.assertIn("INV-2026-1622", invoice_nos)
            self.assertIn("INV-00001", invoice_nos)
            self.assertIn("INV-00002", invoice_nos)

    def test_compute_invoice_rounded_total_uses_half_up_rounding(self):
        self.assertEqual(self.app_module.compute_invoice_rounded_total(48.5), 49.0)
        self.assertEqual(self.app_module.compute_invoice_rounded_total(100.5), 101.0)
        self.assertEqual(self.app_module.compute_invoice_rounded_total("101.49"), 101.0)

    def test_discounted_invoice_print_uses_net_tax_values(self):
        with self.app.app_context():
            self._seed_patient()
            self._seed_vendor_purchase_stack(total_qty=10)

        self.login()
        response = self.client.post(
            "/billing",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "CASH",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["2"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("₹1.19", html)
        self.assertNotIn("₹1.25", html)

    def test_invoice_print_profile_backfill_respects_cutover_date(self):
        with self.app.app_context():
            _patient_old, legacy_invoice = self._seed_invoice(
                customer="Legacy Patient",
                mobile="9000000001",
                created_at=datetime(2026, 6, 30, 12, 0, 0),
            )
            _patient_new, current_invoice = self._seed_invoice(
                customer="Current Patient",
                mobile="9000000002",
                created_at=datetime(2026, 7, 1, 12, 0, 0),
            )

            self.app_module.backfill_invoice_print_profiles()
            self.db.session.refresh(legacy_invoice)
            self.db.session.refresh(current_invoice)

            self.assertEqual(legacy_invoice.print_profile_code, "legacy_pre_2026_07")
            self.assertEqual(legacy_invoice.print_gst_no, "08BPWPP5023C1ZP")
            self.assertEqual(legacy_invoice.print_licence_no, "")

            self.assertEqual(current_invoice.print_profile_code, "endo_pharmacy_2026_07")
            self.assertEqual(current_invoice.print_gst_no, "08ABAFT0637R1ZP")
            self.assertEqual(current_invoice.print_licence_no, "DRUG/2026-27/154505")
            self.assertEqual(
                current_invoice.print_address_line_1,
                "SHOP NO. FF12, 2ND FLOOR, MANGLAM AANANDA PLAZA, SANGANER",
            )

    def test_invoice_print_profile_sync_realigns_existing_snapshots_to_cutover_profiles(self):
        with self.app.app_context():
            _patient_old, legacy_invoice = self._seed_invoice(
                customer="Legacy Snapshot",
                mobile="9000000011",
                created_at=datetime(2026, 6, 30, 12, 0, 0),
            )
            _patient_new, current_invoice = self._seed_invoice(
                customer="Current Snapshot",
                mobile="9000000012",
                created_at=datetime(2026, 7, 2, 12, 0, 0),
            )

            legacy_invoice.print_profile_code = "legacy_pre_2026_07"
            legacy_invoice.print_gst_no = "WRONG-LEGACY-GST"
            current_invoice.print_profile_code = "endo_pharmacy_2026_07"
            current_invoice.print_address_line_1 = "OLD CURRENT ADDRESS"
            current_invoice.print_gst_no = "WRONG-CURRENT-GST"
            self.db.session.commit()

            updated = self.app_module.sync_invoice_print_profiles_to_cutover()
            self.db.session.refresh(legacy_invoice)
            self.db.session.refresh(current_invoice)

            self.assertEqual(updated, 2)
            self.assertEqual(legacy_invoice.print_gst_no, "08BPWPP5023C1ZP")
            self.assertEqual(
                current_invoice.print_address_line_1,
                "SHOP NO. FF12, 2ND FLOOR, MANGLAM AANANDA PLAZA, SANGANER",
            )
            self.assertEqual(current_invoice.print_gst_no, "08ABAFT0637R1ZP")

    def test_invoice_view_prefers_stored_print_snapshot_over_date_fallback(self):
        with self.app.app_context():
            _patient, invoice = self._seed_invoice(
                customer="Snapshot Patient",
                mobile="9000000003",
                created_at=datetime(2026, 7, 2, 12, 0, 0),
            )
            self.app_module.apply_invoice_print_profile(invoice, date(2026, 6, 30))
            self.db.session.commit()
            invoice_id = invoice.id

        self.login()
        response = self.client.get(f"/invoice/{invoice_id}")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("08BPWPP5023C1ZP", html)
        self.assertNotIn("08ABAFT0637R1ZP", html)
        self.assertNotIn("DRUG/2026-27/154505", html)

    def test_invoice_view_shows_local_bill_time_next_to_date(self):
        with self.app.app_context():
            _patient, invoice = self._seed_invoice(
                customer="Time Patient",
                mobile="9000000004",
                created_at=datetime(2026, 7, 2, 12, 0, 0),
            )
            self.app_module.apply_invoice_print_profile(invoice, invoice.created_at)
            self.db.session.commit()
            invoice_id = invoice.id

        self.login()
        response = self.client.get(f"/invoice/{invoice_id}")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("02-07-2026 Time 05:30 PM", html)

    def test_admin_stock_sale_creates_purchase_rate_invoice_and_reentry_vendor_purchase(self):
        with self.app.app_context():
            self._seed_vendor_purchase_stack(total_qty=10, purchase_chunks=[(10, 10.0, datetime.utcnow())])

        self.login()
        response = self.client.post(
            "/stock-sale",
            data={
                "customer": "STOCK SALE",
                "mobile": "",
                "doctor": "Dr. Admin",
                "payment_mode": "ADJUSTMENT",
                "lines_per_invoice": "50",
                "internal_note": "Cycle all stock",
                "confirm_stock_sale": "yes",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Stock sale completed.", response.data)
        self.assertIn(b"DR ABHISHEK PRAKASH", response.data)

        with self.app.app_context():
            medicine = self.app_module.Medicine.query.filter_by(name="PARACETAMOL 650", batch="B123").one()
            vendor = self.app_module.Vendor.query.filter_by(name="DR ABHISHEK PRAKASH").one()
            invoice = self.app_module.Invoice.query.order_by(self.app_module.Invoice.id.desc()).first()
            invoice_item = self.app_module.InvoiceItem.query.filter_by(invoice_id=invoice.id).one()
            reentry_purchase = self.app_module.VendorPurchase.query.filter_by(vendor_id=vendor.id).one()
            reentry_item = self.app_module.VendorPurchaseItem.query.filter_by(purchase_id=reentry_purchase.id).one()
            total_remaining = sum(
                item.remaining_qty
                for item in self.app_module.VendorPurchaseItem.query.filter_by(
                    medicine_name="PARACETAMOL 650",
                    batch="B123",
                ).all()
            )
            sale_history_count = self.app_module.StockHistory.query.filter(
                self.app_module.StockHistory.remark.like("Bulk stock sale%")
            ).count()
            purchase_history_count = self.app_module.StockHistory.query.filter(
                self.app_module.StockHistory.remark.like("Auto re-entry%")
            ).count()

            self.assertEqual(medicine.qty, 10)
            self.assertEqual(invoice.customer, "STOCK SALE")
            self.assertEqual(invoice.payment_mode, "ADJUSTMENT")
            self.assertIn("Bulk stock sale operation", invoice.internal_note)
            self.assertEqual(float(invoice_item.price), 10.0)
            self.assertEqual(float(invoice_item.net_amount), 100.0)
            self.assertEqual(reentry_purchase.payment_status, "Paid")
            self.assertEqual(reentry_item.remaining_qty, 10)
            self.assertEqual(float(reentry_item.purchase_rate), 10.0)
            self.assertEqual(total_remaining, 10)
            self.assertEqual(sale_history_count, 1)
            self.assertEqual(purchase_history_count, 1)

    def test_stock_sale_supports_credit_payment_mode_and_party_gst_on_invoice(self):
        with self.app.app_context():
            self._seed_vendor_purchase_stack(total_qty=6, purchase_chunks=[(6, 10.0, datetime.utcnow())])

        self.login()
        response = self.client.post(
            "/stock-sale",
            data={
                "customer": "M/S SHARMA TRADERS",
                "customer_gst_no": "08ABCDE1234F1Z5",
                "mobile": "9000000001",
                "doctor": "Dr. Admin",
                "payment_mode": "CREDIT",
                "lines_per_invoice": "50",
                "confirm_stock_sale": "yes",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Stock sale completed.", response.data)

        with self.app.app_context():
            invoice = self.app_module.Invoice.query.order_by(self.app_module.Invoice.id.desc()).first()
            self.assertEqual(invoice.payment_mode, "CREDIT")
            self.assertEqual(float(invoice.cash_amount or 0), 0.0)
            self.assertEqual(float(invoice.online_amount or 0), 0.0)
            self.assertEqual(invoice.customer_gst_no, "08ABCDE1234F1Z5")
            invoice_id = invoice.id

        response = self.client.get(f"/invoice/{invoice_id}")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("Party Details", html)
        self.assertIn("Party Name", html)
        self.assertIn("GST No", html)
        self.assertIn("08ABCDE1234F1Z5", html)
        self.assertIn("CREDIT", html)

    def test_stock_sale_preview_uses_medicine_qty_and_clears_stale_purchase_layer_surplus(self):
        with self.app.app_context():
            _vendor, _purchase, medicine = self._seed_vendor_purchase_stack(
                total_qty=10,
                purchase_chunks=[
                    (5, 10.0, datetime.utcnow() - timedelta(days=2)),
                    (5, 12.0, datetime.utcnow() - timedelta(days=1)),
                ],
            )
            medicine.qty = 8
            self.db.session.commit()

            preview = self.app_module.build_stock_sale_preview(50)
            self.assertEqual(preview["estimated_sale_value"], 86.0)
            self.assertEqual(preview["line_count"], 2)
            self.assertEqual(preview["total_qty"], 8)
            self.assertEqual(preview["sync_mismatch_count"], 0)
            self.assertEqual(preview["sync_total_delta"], 0)
            self.assertEqual(preview["ignored_layer_qty"], 2)
            self.assertEqual(preview["ignored_layer_count"], 1)

        self.login()
        response = self.client.post(
            "/stock-sale",
            data={
                "customer": "STOCK SALE",
                "payment_mode": "ADJUSTMENT",
                "lines_per_invoice": "50",
                "confirm_stock_sale": "yes",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Stock sale completed.", response.data)

        with self.app.app_context():
            medicine = self.app_module.Medicine.query.filter_by(name="PARACETAMOL 650", batch="B123").one()
            adjustment_history = self.app_module.StockHistory.query.filter(
                self.app_module.StockHistory.remark.like("Stock sale sync%")
            ).all()
            total_remaining = sum(
                item.remaining_qty
                for item in self.app_module.VendorPurchaseItem.query.filter_by(
                    medicine_name="PARACETAMOL 650",
                    batch="B123",
                ).all()
            )
            reentry_purchase = self.app_module.VendorPurchase.query.order_by(
                self.app_module.VendorPurchase.id.desc()
            ).first()
            reentry_quantities = sorted(
                item.remaining_qty
                for item in self.app_module.VendorPurchaseItem.query.filter_by(
                    purchase_id=reentry_purchase.id
                ).all()
            )
            self.assertEqual(medicine.qty, 8)
            self.assertEqual(adjustment_history, [])
            self.assertEqual(total_remaining, 8)
            self.assertEqual(reentry_quantities, [3, 5])

    def test_stock_sale_preview_ignores_zero_qty_medicine_even_if_old_purchase_layer_exists(self):
        with self.app.app_context():
            _vendor, _purchase, medicine = self._seed_vendor_purchase_stack(total_qty=10)
            medicine.qty = 0
            self.db.session.commit()

            preview = self.app_module.build_stock_sale_preview(50)
            self.assertEqual(preview["line_count"], 0)
            self.assertEqual(preview["total_qty"], 0)
            self.assertEqual(preview["ignored_layer_qty"], 10)
            self.assertEqual(preview["ignored_layer_count"], 1)

    def test_stock_sale_is_admin_only_even_for_staff_with_purchase_access(self):
        with self.app.app_context():
            staff = self.app_module.User(
                username="stockstaff",
                role="staff",
                access_profile="custom",
                can_manage_purchases=True,
                can_invoice_action=True,
                can_edit_invoice=True,
            )
            staff.set_password("Staff@123")
            self.db.session.add(staff)
            self.db.session.commit()

        self.login_as("stockstaff", "Staff@123")
        response = self.client.get("/stock-sale", follow_redirects=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Access denied", response.data)

    def test_stock_sale_blocks_execution_when_purchase_rate_history_is_missing(self):
        with self.app.app_context():
            self.db.session.add(
                self.app_module.Medicine(
                    name="UNMAPPED STOCK",
                    batch="U-100",
                    expiry="2027-12-31",
                    mrp=55.0,
                    qty=4,
                    discount_percent=0,
                    barcode="UNMAPPED100",
                    reorder_level=1,
                    is_active=True,
                )
            )
            self.db.session.commit()

        self.login()
        response = self.client.post(
            "/stock-sale",
            data={
                "customer": "STOCK SALE",
                "payment_mode": "ADJUSTMENT",
                "lines_per_invoice": "50",
                "confirm_stock_sale": "yes",
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"no purchase layer", response.data.lower())

        with self.app.app_context():
            self.assertEqual(self.app_module.Invoice.query.count(), 0)

    def test_billing_rejects_split_cash_above_total_bill_amount(self):
        with self.app.app_context():
            self._seed_patient()
            self._seed_vendor_purchase_stack(total_qty=10)

        self.login()
        response = self.client.post(
            "/billing",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "UPI",
                "split_cash_amount": "80.00",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["2"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Cash amount cannot exceed total bill amount.", response.data)

        with self.app.app_context():
            self.assertEqual(self.app_module.Invoice.query.count(), 0)

    def test_vendor_purchase_syncs_barcode_to_medicine_master(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="Sync Vendor", is_active=True)
            medicine = self.app_module.Medicine(
                name="CALCITAB",
                medicine_code="MED-CALCI",
                batch="C100",
                expiry="2027-12-31",
                mrp=120.0,
                qty=2,
                discount_percent=10,
                barcode="",
                reorder_level=5,
                is_active=True,
            )
            self.db.session.add(vendor)
            self.db.session.add(medicine)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "SYNC-001",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": ["CALCITAB"],
                "medicine_code": ["MED-CALCI"],
                "barcode": ["BARCODE-12345"],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Sync Vendor"],
                "pack_type": ["Box"],
                "pack_qty": ["1"],
                "batch": ["C100"],
                "expiry": ["12/2027"],
                "qty": ["5"],
                "free_qty": ["0"],
                "purchase_rate": ["80"],
                "mrp": ["120"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            expected_code = self._expected_auto_code("CALCITAB")
            medicine = self.app_module.Medicine.query.filter_by(name="CALCITAB", batch="C100").one()
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            self.assertEqual(medicine.medicine_code, expected_code)
            self.assertEqual(medicine.barcode, "BARCODE-12345")
            self.assertEqual(purchase_item.medicine_code, expected_code)
            self.assertEqual(purchase_item.barcode, "BARCODE-12345")

    def test_vendor_purchase_resolves_existing_medicine_name_from_code(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="Code Vendor", is_active=True)
            medicine = self.app_module.Medicine(
                name="CALCITAB D3 LIQUID 5 ML 1X1",
                medicine_code="MED-D3LQ",
                batch="OLD100",
                expiry="2027-12-31",
                mrp=140.0,
                qty=4,
                discount_percent=10,
                barcode="",
                reorder_level=5,
                is_active=True,
            )
            self.db.session.add(vendor)
            self.db.session.add(medicine)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "SYNC-002",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": [""],
                "medicine_code": ["MED-D3LQ"],
                "barcode": [""],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Code Vendor"],
                "pack_type": ["Bottle"],
                "pack_qty": ["1"],
                "batch": ["NEW200"],
                "expiry": ["12/2027"],
                "qty": ["5"],
                "free_qty": ["0"],
                "purchase_rate": ["80"],
                "mrp": ["140"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            expected_code = self._expected_auto_code("CALCITAB D3 LIQUID 5 ML 1X1")
            new_batch = self.app_module.Medicine.query.filter_by(batch="NEW200").one()
            purchase_item = self.app_module.VendorPurchaseItem.query.order_by(
                self.app_module.VendorPurchaseItem.id.desc()
            ).first()
            self.assertEqual(new_batch.name, "CALCITAB D3 LIQUID 5 ML 1X1")
            self.assertEqual(new_batch.medicine_code, expected_code)
            self.assertEqual(purchase_item.medicine_name, "CALCITAB D3 LIQUID 5 ML 1X1")
            self.assertEqual(purchase_item.medicine_code, expected_code)

    def test_vendor_purchase_accepts_ajax_json_purchase_grid_payload(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="ERP Vendor", is_active=True)
            medicine = self.app_module.Medicine(
                name="DOLO 650",
                medicine_code="MED-DOLO650",
                batch="OLD-1",
                expiry="2027-12-31",
                mrp=35.0,
                qty=10,
                discount_percent=5,
                barcode="8900000000011",
                reorder_level=5,
                is_active=True,
                composition="Paracetamol 650",
                company="Micro Labs",
                pack_type="Strip",
                pack_qty=15,
            )
            self.db.session.add(vendor)
            self.db.session.add(medicine)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        payload = [
            {
                "medicine_code": "MED-DOLO650",
                "medicine_name": "DOLO 650",
                "barcode": "8900000000011",
                "batch": "NEW-1",
                "expiry": "12/2027",
                "pack_type": "Strip",
                "pack_qty": 15,
                "qty": 20,
                "free_qty": 2,
                "purchase_rate": 18,
                "mrp": 35,
                "gst_percent": 12,
                "discount_percent": 5,
                "composition": "Paracetamol 650",
                "company": "Micro Labs",
                "distributor_name": "ERP Vendor",
                "notes": "Fast lane item",
            }
        ]
        response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "ERP-001",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "purchase_notes": "Main purchase note",
                "purchase_items_json": self.app_module.json.dumps(payload),
            },
            headers={
                "X-Requested-With": "XMLHttpRequest",
                "Accept": "application/json",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        body = response.get_json()
        self.assertTrue(body["ok"])
        self.assertEqual(body["purchase_no"][:3], "PB-")

        with self.app.app_context():
            purchase = self.app_module.VendorPurchase.query.one()
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            saved_medicine = self.app_module.Medicine.query.filter_by(batch="NEW-1").one()
            expected_code = self._expected_auto_code("DOLO 650")
            self.assertEqual(purchase.notes, "Main purchase note")
            self.assertEqual(purchase_item.notes, "Fast lane item")
            self.assertEqual(saved_medicine.medicine_code, expected_code)
            self.assertEqual(purchase_item.medicine_code, expected_code)
            self.assertEqual(saved_medicine.qty, 22)

    def test_vendor_purchase_starts_fresh_start_purchase_series_from_one(self):
        with self.app.app_context():
            vendor, _purchase, _medicine = self._seed_vendor_purchase_stack(total_qty=10)
            vendor_id = vendor.id

        self.login()
        first_response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "FRESH-PUR-001",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": ["PARACETAMOL 650"],
                "medicine_code": [""],
                "barcode": [""],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Prime Distributor"],
                "pack_type": ["Box"],
                "pack_qty": ["1"],
                "batch": ["P001"],
                "expiry": ["12/2027"],
                "qty": ["3"],
                "free_qty": ["0"],
                "purchase_rate": ["15"],
                "mrp": ["25"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(first_response.status_code, 302)

        second_response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "FRESH-PUR-002",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": ["PARACETAMOL 650"],
                "medicine_code": [""],
                "barcode": [""],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Prime Distributor"],
                "pack_type": ["Box"],
                "pack_qty": ["1"],
                "batch": ["P002"],
                "expiry": ["12/2027"],
                "qty": ["2"],
                "free_qty": ["0"],
                "purchase_rate": ["16"],
                "mrp": ["25"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(second_response.status_code, 302)

        with self.app.app_context():
            purchase_nos = [
                row[0]
                for row in self.db.session.query(self.app_module.VendorPurchase.purchase_no)
                .order_by(self.app_module.VendorPurchase.id.asc())
                .all()
            ]
            self.assertIn("PB-000001", purchase_nos)
            self.assertIn("PB-00001", purchase_nos)
            self.assertIn("PB-00002", purchase_nos)

    def test_vendor_form_renders_auto_code_purchase_builder_for_manual_medicine_entry(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="Readonly Vendor", is_active=True)
            self.db.session.add(vendor)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        response = self.client.get(f"/vendor/edit/{vendor_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'purchase_items_json', response.data)
        self.assertIn(b'Add Medicine', response.data)
        self.assertIn(b'Live Purchase Grid', response.data)
        self.assertIn(b'id="purchaseModal"', response.data)
        self.assertIn(b'Auto-generated on save', response.data)
        self.assertIn(b'Type or confirm medicine name', response.data)
        self.assertIn(b"purchase_items_json", response.data)

    def test_vendor_bill_history_uses_three_dot_action_menu_and_keeps_existing_routes(self):
        with self.app.app_context():
            vendor, purchase, _medicine = self._seed_vendor_purchase_stack(total_qty=6)
            vendor_id = vendor.id
            purchase_id = purchase.id

        self.login()
        response = self.client.get(f"/vendor/edit/{vendor_id}")
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Vendor Bill History", response.data)
        self.assertIn(b'vendor-history-action-menu', response.data)
        self.assertIn(b'aria-label="More Actions"', response.data)
        self.assertIn(f'/vendor/purchase/{purchase_id}'.encode(), response.data)
        self.assertIn(f'/vendor/purchase/{purchase_id}?mode=edit'.encode(), response.data)
        self.assertIn(f'/vendor/purchase/{purchase_id}?mode=return'.encode(), response.data)
        self.assertIn(f'/vendor/purchase/delete/{purchase_id}'.encode(), response.data)

    def test_vendor_purchase_edit_mode_allows_deleting_unsold_item_and_updates_stock(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(
                name="Delete Item Vendor",
                is_active=True,
                total_purchases=130.0,
                payment_status="Paid",
            )
            self.db.session.add(vendor)
            self.db.session.flush()

            medicine_a = self.app_module.Medicine(
                name="ITEM A TAB",
                medicine_code="ITEMA001",
                batch="A1",
                expiry="2027-12-31",
                mrp=12.0,
                qty=5,
                discount_percent=0,
                is_active=True,
            )
            medicine_b = self.app_module.Medicine(
                name="ITEM B TAB",
                medicine_code="ITEMB001",
                batch="B1",
                expiry="2027-12-31",
                mrp=25.0,
                qty=4,
                discount_percent=0,
                is_active=True,
            )
            self.db.session.add_all([medicine_a, medicine_b])
            self.db.session.flush()

            purchase = self.app_module.VendorPurchase(
                vendor_id=vendor.id,
                purchase_no="PB-DEL-001",
                invoice_no="SUP-DEL-001",
                purchase_date=datetime.utcnow(),
                payment_mode="CASH",
                payment_status="Paid",
                paid_amount=130.0,
                subtotal=130.0,
                gst_total=0.0,
                discount_total=0.0,
                total_amount=130.0,
                created_by="admin",
            )
            self.db.session.add(purchase)
            self.db.session.flush()

            item_a = self.app_module.VendorPurchaseItem(
                purchase_id=purchase.id,
                vendor_id=vendor.id,
                medicine_id=medicine_a.id,
                medicine_name=medicine_a.name,
                medicine_code=medicine_a.medicine_code,
                batch=medicine_a.batch,
                expiry=medicine_a.expiry,
                qty=5,
                free_qty=0,
                remaining_qty=5,
                purchase_rate=10.0,
                mrp=medicine_a.mrp,
                gst_percent=0.0,
                discount_percent=0.0,
                total_value=50.0,
            )
            item_b = self.app_module.VendorPurchaseItem(
                purchase_id=purchase.id,
                vendor_id=vendor.id,
                medicine_id=medicine_b.id,
                medicine_name=medicine_b.name,
                medicine_code=medicine_b.medicine_code,
                batch=medicine_b.batch,
                expiry=medicine_b.expiry,
                qty=4,
                free_qty=0,
                remaining_qty=4,
                purchase_rate=20.0,
                mrp=medicine_b.mrp,
                gst_percent=0.0,
                discount_percent=0.0,
                total_value=80.0,
            )
            self.db.session.add_all([item_a, item_b])
            self.db.session.commit()
            purchase_id = purchase.id
            item_a_id = item_a.id
            medicine_a_id = medicine_a.id
            vendor_id = vendor.id

        self.login()
        response = self.client.get(f"/vendor/purchase/{purchase_id}?mode=edit")
        self.assertEqual(response.status_code, 200)
        self.assertIn(f'/vendor/purchase-item/delete/{item_a_id}'.encode(), response.data)
        self.assertIn(b"Delete Item", response.data)

        response = self.client.post(
            f"/vendor/purchase-item/delete/{item_a_id}",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(f"/vendor/purchase/{purchase_id}?mode=edit", response.headers["Location"])

        with self.app.app_context():
            remaining_items = self.app_module.VendorPurchaseItem.query.filter_by(purchase_id=purchase_id).all()
            medicine_a = self.app_module.Medicine.query.get(medicine_a_id)
            vendor = self.app_module.Vendor.query.get(vendor_id)
            purchase = self.app_module.VendorPurchase.query.get(purchase_id)
            delete_history = self.app_module.StockHistory.query.filter_by(action="PURCHASE_ITEM_DELETE").one()

            self.assertEqual(len(remaining_items), 1)
            self.assertEqual(remaining_items[0].medicine_name, "ITEM B TAB")
            self.assertEqual(medicine_a.qty, 0)
            self.assertEqual(float(purchase.total_amount), 80.0)
            self.assertEqual(float(purchase.subtotal), 80.0)
            self.assertEqual(float(vendor.total_purchases), 80.0)
            self.assertEqual(delete_history.qty_change, -5)

    def test_vendor_purchase_item_edit_returns_json_and_stays_on_same_page(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(
                name="Edit Item Vendor",
                is_active=True,
                total_purchases=50.0,
                payment_status="Paid",
            )
            self.db.session.add(vendor)
            self.db.session.flush()

            medicine = self.app_module.Medicine(
                name="EDIT ITEM TAB",
                medicine_code="EDIT001",
                batch="E1",
                expiry="2027-12-31",
                mrp=12.0,
                qty=5,
                discount_percent=0,
                is_active=True,
            )
            self.db.session.add(medicine)
            self.db.session.flush()

            purchase = self.app_module.VendorPurchase(
                vendor_id=vendor.id,
                purchase_no="PB-EDIT-001",
                invoice_no="SUP-EDIT-001",
                purchase_date=datetime.utcnow(),
                payment_mode="CASH",
                payment_status="Paid",
                paid_amount=50.0,
                subtotal=50.0,
                gst_total=0.0,
                discount_total=0.0,
                total_amount=50.0,
            )
            self.db.session.add(purchase)
            self.db.session.flush()

            item = self.app_module.VendorPurchaseItem(
                purchase_id=purchase.id,
                vendor_id=vendor.id,
                medicine_id=medicine.id,
                medicine_name=medicine.name,
                medicine_code=medicine.medicine_code,
                batch=medicine.batch,
                expiry=medicine.expiry,
                qty=5,
                free_qty=0,
                remaining_qty=5,
                purchase_rate=10.0,
                mrp=medicine.mrp,
                gst_percent=0.0,
                discount_percent=0.0,
                total_value=50.0,
            )
            self.db.session.add(item)
            self.db.session.commit()
            item_id = item.id
            medicine_id = medicine.id
            purchase_id = purchase.id

        self.login()
        response = self.client.post(
            f"/vendor/purchase-item/edit/{item_id}",
            data={
                "medicine_name": "EDIT ITEM TAB UPDATED",
                "medicine_code": "EDIT001",
                "barcode": "",
                "composition": "",
                "company": "",
                "distributor_name": "",
                "pack_type": "Strip",
                "pack_qty": "10",
                "batch": "E2",
                "expiry": "12/2027",
                "qty": "7",
                "free_qty": "1",
                "purchase_rate": "10",
                "mrp": "12",
                "gst_percent": "0",
                "discount_percent": "0",
            },
            headers={"X-Requested-With": "XMLHttpRequest", "Accept": "application/json"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["message"], "Purchase item updated successfully")
        self.assertEqual(payload["item"]["medicine_name"], "EDIT ITEM TAB UPDATED")
        self.assertEqual(payload["item"]["batch"], "E2")
        self.assertEqual(payload["item"]["qty"], 7)
        self.assertEqual(payload["item"]["free_qty"], 1)
        self.assertEqual(payload["purchase_item_count"], 1)
        self.assertEqual(float(payload["purchase_totals"]["total_amount"]), 70.0)

        with self.app.app_context():
            purchase = self.app_module.VendorPurchase.query.get(purchase_id)
            medicine = self.app_module.Medicine.query.get(medicine_id)
            item = self.app_module.VendorPurchaseItem.query.get(item_id)
            self.assertEqual(item.medicine_name, "EDIT ITEM TAB UPDATED")
            self.assertEqual(item.batch, "E2")
            self.assertEqual(item.qty, 7)
            self.assertEqual(item.free_qty, 1)
            self.assertEqual(item.remaining_qty, 8)
            self.assertEqual(medicine.name, "EDIT ITEM TAB UPDATED")
            self.assertEqual(medicine.batch, "E2")
            self.assertEqual(medicine.qty, 8)
            self.assertEqual(float(purchase.total_amount), 70.0)

    def test_vendor_purchase_item_delete_returns_json_and_stays_on_same_page(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(
                name="Delete Ajax Vendor",
                is_active=True,
                total_purchases=130.0,
                payment_status="Paid",
            )
            self.db.session.add(vendor)
            self.db.session.flush()

            medicine_a = self.app_module.Medicine(
                name="AJAX ITEM A TAB",
                medicine_code="AJA001",
                batch="AA1",
                expiry="2027-12-31",
                mrp=12.0,
                qty=5,
                discount_percent=0,
                is_active=True,
            )
            medicine_b = self.app_module.Medicine(
                name="AJAX ITEM B TAB",
                medicine_code="AJB001",
                batch="BB1",
                expiry="2027-12-31",
                mrp=25.0,
                qty=4,
                discount_percent=0,
                is_active=True,
            )
            self.db.session.add_all([medicine_a, medicine_b])
            self.db.session.flush()

            purchase = self.app_module.VendorPurchase(
                vendor_id=vendor.id,
                purchase_no="PB-AJAX-001",
                invoice_no="SUP-AJAX-001",
                purchase_date=datetime.utcnow(),
                payment_mode="CASH",
                payment_status="Paid",
                paid_amount=130.0,
                subtotal=130.0,
                gst_total=0.0,
                discount_total=0.0,
                total_amount=130.0,
            )
            self.db.session.add(purchase)
            self.db.session.flush()

            item_a = self.app_module.VendorPurchaseItem(
                purchase_id=purchase.id,
                vendor_id=vendor.id,
                medicine_id=medicine_a.id,
                medicine_name=medicine_a.name,
                medicine_code=medicine_a.medicine_code,
                batch=medicine_a.batch,
                expiry=medicine_a.expiry,
                qty=5,
                free_qty=0,
                remaining_qty=5,
                purchase_rate=10.0,
                mrp=medicine_a.mrp,
                gst_percent=0.0,
                discount_percent=0.0,
                total_value=50.0,
            )
            item_b = self.app_module.VendorPurchaseItem(
                purchase_id=purchase.id,
                vendor_id=vendor.id,
                medicine_id=medicine_b.id,
                medicine_name=medicine_b.name,
                medicine_code=medicine_b.medicine_code,
                batch=medicine_b.batch,
                expiry=medicine_b.expiry,
                qty=4,
                free_qty=0,
                remaining_qty=4,
                purchase_rate=20.0,
                mrp=medicine_b.mrp,
                gst_percent=0.0,
                discount_percent=0.0,
                total_value=80.0,
            )
            self.db.session.add_all([item_a, item_b])
            self.db.session.commit()
            item_a_id = item_a.id
            purchase_id = purchase.id
            medicine_a_id = medicine_a.id

        self.login()
        response = self.client.post(
            f"/vendor/purchase-item/delete/{item_a_id}",
            headers={"X-Requested-With": "XMLHttpRequest", "Accept": "application/json"},
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["deleted_item_id"], item_a_id)
        self.assertEqual(payload["purchase_item_count"], 1)
        self.assertEqual(float(payload["purchase_totals"]["total_amount"]), 80.0)

        with self.app.app_context():
            remaining_items = self.app_module.VendorPurchaseItem.query.filter_by(purchase_id=purchase_id).all()
            medicine_a = self.app_module.Medicine.query.get(medicine_a_id)
            self.assertEqual(len(remaining_items), 1)
            self.assertEqual(remaining_items[0].medicine_name, "AJAX ITEM B TAB")
            self.assertEqual(medicine_a.qty, 0)

    def test_vendor_form_hides_address_bank_and_payment_sections_but_preserves_existing_values(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(
                name="Preserve Vendor",
                mobile="9999988888",
                email="vendor@example.com",
                gst_no="GST123",
                shop_name="Old Shop",
                area="Old Area",
                city="Old City",
                state="Old State",
                pincode="123456",
                address="Old Address",
                bank_name="Old Bank",
                account_holder_name="Old Holder",
                account_no="1234567890",
                ifsc="IFSC0001",
                upi="old@upi",
                vendor_type="Distributor",
                default_payment_mode="CASH",
                payment_status="Paid",
                attachment_ref="old-file.pdf",
                notes="Old vendor note",
                is_active=True,
            )
            self.db.session.add(vendor)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        get_response = self.client.get(f"/vendor/edit/{vendor_id}")
        self.assertEqual(get_response.status_code, 200)
        vendor_card_html = get_response.data.split(b'<div class="purchase-workspace">', 1)[0]
        self.assertNotIn(b"Address", vendor_card_html)
        self.assertNotIn(b"Bank Details", vendor_card_html)
        self.assertNotIn(b"Shop Name", vendor_card_html)
        self.assertNotIn(b"Account Holder Name", vendor_card_html)
        self.assertNotIn(b'name="vendor_type"', vendor_card_html)
        self.assertNotIn(b'name="default_payment_mode"', vendor_card_html)
        self.assertNotIn(b'name="payment_status"', vendor_card_html)
        self.assertNotIn(b'name="last_purchase_date"', vendor_card_html)
        self.assertNotIn(b'name="is_active"', vendor_card_html)
        self.assertNotIn(b"Attachments & Notes", vendor_card_html)
        self.assertNotIn(b'name="attachment_ref"', vendor_card_html)
        self.assertNotIn(b'name="attachment_file"', vendor_card_html)
        self.assertNotIn(b'name="notes"', vendor_card_html)

        post_response = self.client.post(
            f"/vendor/edit/{vendor_id}",
            data={
                "name": "Preserve Vendor Updated",
                "mobile": "9999988888",
                "email": "vendor@example.com",
                "gst_no": "GST123",
            },
            follow_redirects=False,
        )
        self.assertEqual(post_response.status_code, 302)

        with self.app.app_context():
            vendor = self.app_module.Vendor.query.get(vendor_id)
            self.assertEqual(vendor.shop_name, "Old Shop")
            self.assertEqual(vendor.area, "Old Area")
            self.assertEqual(vendor.city, "Old City")
            self.assertEqual(vendor.state, "Old State")
            self.assertEqual(vendor.pincode, "123456")
            self.assertEqual(vendor.address, "Old Address")
            self.assertEqual(vendor.bank_name, "Old Bank")
            self.assertEqual(vendor.account_holder_name, "Old Holder")
            self.assertEqual(vendor.account_no, "1234567890")
            self.assertEqual(vendor.ifsc, "IFSC0001")
            self.assertEqual(vendor.upi, "old@upi")
            self.assertEqual(vendor.vendor_type, "Distributor")
            self.assertEqual(vendor.default_payment_mode, "CASH")
            self.assertEqual(vendor.payment_status, "Paid")
            self.assertEqual(vendor.attachment_ref, "old-file.pdf")
            self.assertEqual(vendor.notes, "Old vendor note")
            self.assertTrue(vendor.is_active)

    def test_vendor_purchase_auto_generates_code_for_manual_or_unknown_entries(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="Strict Code Vendor", is_active=True)
            self.db.session.add(vendor)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        missing_code_response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "STRICT-001",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": ["MANUAL MED"],
                "medicine_code": [""],
                "barcode": [""],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Strict Code Vendor"],
                "pack_type": ["Box"],
                "pack_qty": ["1"],
                "batch": ["MC100"],
                "expiry": ["12/2027"],
                "qty": ["2"],
                "free_qty": ["0"],
                "purchase_rate": ["50"],
                "mrp": ["70"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(missing_code_response.status_code, 302)

        unknown_code_response = self.client.post(
            f"/vendor/{vendor_id}/purchase",
            data={
                "invoice_no": "STRICT-002",
                "purchase_date": date.today().isoformat(),
                "payment_mode": "CASH",
                "payment_status": "Paid",
                "paid_amount": "0",
                "medicine_name": ["MYSTERY MED"],
                "medicine_code": ["MED-UNKNOWN"],
                "barcode": [""],
                "composition": [""],
                "company": ["Test Pharma"],
                "distributor_name": ["Strict Code Vendor"],
                "pack_type": ["Box"],
                "pack_qty": ["1"],
                "batch": ["MC101"],
                "expiry": ["12/2027"],
                "qty": ["2"],
                "free_qty": ["0"],
                "purchase_rate": ["50"],
                "mrp": ["70"],
                "gst_percent": ["0"],
                "discount_percent": ["0"],
            },
            follow_redirects=False,
        )
        self.assertEqual(unknown_code_response.status_code, 302)

        with self.app.app_context():
            self.assertEqual(self.app_module.VendorPurchase.query.count(), 2)
            generated_names = {
                med.name: med.medicine_code
                for med in self.app_module.Medicine.query.order_by(self.app_module.Medicine.name.asc()).all()
            }
            self.assertEqual(generated_names["MANUAL MED"], self._expected_auto_code("MANUAL MED"))
            self.assertEqual(generated_names["MYSTERY MED"], self._expected_auto_code("MYSTERY MED"))

    def test_vendor_purchase_succeeds_with_legacy_integer_medicine_is_active_storage(self):
        with self.app.app_context():
            vendor = self.app_module.Vendor(name="Legacy Storage Vendor", is_active=True)
            self.db.session.add(vendor)
            self.db.session.commit()
            vendor_id = vendor.id

        self.login()
        original_runtime_boolean_storage_mode = self.app_module.runtime_boolean_storage_mode

        def fake_runtime_boolean_storage_mode(table_name, column_name):
            if table_name == "medicine" and column_name == "is_active":
                return "integer"
            return original_runtime_boolean_storage_mode(table_name, column_name)

        self.app_module.runtime_boolean_storage_mode = fake_runtime_boolean_storage_mode
        try:
            response = self.client.post(
                f"/vendor/{vendor_id}/purchase",
                data={
                    "invoice_no": "LEGACY-INT-001",
                    "purchase_date": date.today().isoformat(),
                    "payment_mode": "CASH",
                    "payment_status": "Paid",
                    "paid_amount": "0",
                    "medicine_name": ["SUNDAE 2 MG/1.5 ML KWIK PEN"],
                    "medicine_code": [""],
                    "barcode": [""],
                    "composition": [""],
                    "company": ["Legacy Pharma"],
                    "distributor_name": ["Legacy Storage Vendor"],
                    "pack_type": ["Pen"],
                    "pack_qty": ["1"],
                    "batch": ["7000631B"],
                    "expiry": ["02/2028"],
                    "qty": ["2"],
                    "free_qty": ["0"],
                    "purchase_rate": ["2216.46"],
                    "mrp": ["3200"],
                    "gst_percent": ["5"],
                    "discount_percent": ["5"],
                },
                follow_redirects=False,
            )
        finally:
            self.app_module.runtime_boolean_storage_mode = original_runtime_boolean_storage_mode

        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            med = self.app_module.Medicine.query.filter_by(batch="7000631B").one()
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            self.assertEqual(med.medicine_code, self._expected_auto_code("SUNDAE 2 MG/1.5 ML KWIK PEN"))
            self.assertEqual(med.qty, 2)
            self.assertEqual(purchase_item.medicine_code, med.medicine_code)

    def test_sync_medicine_codes_to_current_names_replaces_manual_codes_across_batches(self):
        with self.app.app_context():
            med_one = self.app_module.Medicine(
                name="DOLO 650",
                medicine_code="MED0001",
                batch="D-1",
                expiry="2027-12-31",
                mrp=35.0,
                qty=5,
                discount_percent=5,
                is_active=True,
            )
            med_two = self.app_module.Medicine(
                name="DOLO 650",
                medicine_code="MED0099",
                batch="D-2",
                expiry="2027-12-31",
                mrp=35.0,
                qty=8,
                discount_percent=5,
                is_active=True,
            )
            self.db.session.add_all([med_one, med_two])
            self.db.session.flush()
            purchase_item = self.app_module.VendorPurchaseItem(
                purchase_id=1,
                vendor_id=1,
                medicine_id=med_one.id,
                medicine_name="DOLO 650",
                medicine_code="MED0001",
                batch="D-1",
                expiry="12/2027",
                qty=2,
                free_qty=0,
                remaining_qty=2,
                purchase_rate=20,
                mrp=35,
                gst_percent=0,
                discount_percent=0,
                total_value=40,
            )
            self.db.session.add(purchase_item)
            self.db.session.commit()

            sync_result = self.app_module.sync_medicine_codes_to_current_names()
            self.db.session.commit()

            refreshed_one = self.app_module.Medicine.query.filter_by(batch="D-1").one()
            refreshed_two = self.app_module.Medicine.query.filter_by(batch="D-2").one()
            refreshed_item = self.app_module.VendorPurchaseItem.query.one()
            expected_code = self._expected_auto_code("DOLO 650")

        self.assertGreaterEqual(sync_result["medicine_updates"], 2)
        self.assertEqual(refreshed_one.medicine_code, expected_code)
        self.assertEqual(refreshed_two.medicine_code, expected_code)
        self.assertEqual(refreshed_item.medicine_code, expected_code)

    def test_add_medicine_auto_generates_code_from_current_name(self):
        self.login()
        response = self.client.post(
            "/medicines/add",
            data={
                "name": "CALPOL 650",
                "medicine_code": "MED9999",
                "batch": "CP-1",
                "barcode": "",
                "expiry": "2027-12-31",
                "pack_type": "Strip",
                "pack_qty": "15",
                "mrp": "32",
                "discount_percent": "0",
                "reorder_level": "10",
                "qty": "6",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            med = self.app_module.Medicine.query.filter_by(name="CALPOL 650", batch="CP-1").one()
            self.assertEqual(med.medicine_code, self._expected_auto_code("CALPOL 650"))

    def test_medicines_route_persists_existing_manual_codes_to_name_based_codes(self):
        with self.app.app_context():
            med = self.app_module.Medicine(
                name="CIPLAR 10 MG TAB 1*15",
                medicine_code="MED0018",
                batch="CIP-1",
                expiry="2028-07-31",
                mrp=45.0,
                qty=6,
                discount_percent=0,
                is_active=True,
            )
            self.db.session.add(med)
            self.db.session.commit()

        self.login()
        response = self.client.get("/medicines")
        self.assertEqual(response.status_code, 200)

        with self.app.app_context():
            med = self.app_module.Medicine.query.filter_by(batch="CIP-1").one()
            self.assertEqual(med.medicine_code, self._expected_auto_code("CIPLAR 10 MG TAB 1*15"))

    def test_return_flow_restores_stock_and_purchase_remaining(self):
        with self.app.app_context():
            patient = self._seed_patient()
            vendor, purchase, medicine = self._seed_vendor_purchase_stack(total_qty=3)
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            purchase_item.remaining_qty = 0
            medicine.qty = 0
            invoice = self.app_module.Invoice(
                invoice_no="INV-2001",
                patient_id=patient.id,
                customer=patient.name,
                mobile=patient.mobile,
                subtotal=75.0,
                total=75.0,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(invoice)
            self.db.session.flush()
            invoice_item = self.app_module.InvoiceItem(
                invoice_id=invoice.id,
                name=medicine.name,
                qty=3,
                price=25.0,
                amount=75.0,
                batch=medicine.batch,
                expiry=medicine.expiry,
                discount_percent=0,
                discount_amount=0,
                net_amount=75.0,
                cost_price=10.0,
                cost_amount=30.0,
            )
            self.db.session.add(invoice_item)
            self.db.session.flush()
            self.db.session.add(
                self.app_module.SalesAllocation(
                    invoice_item_id=invoice_item.id,
                    purchase_item_id=purchase_item.id,
                    qty=3,
                    cost_rate=10.0,
                    returned_qty=0,
                )
            )
            self.db.session.commit()
            invoice_item_id = invoice_item.id

        self.login()
        response = self.client.post(
            "/return-medicine",
            data={
                "invoice_no": "INV-2001",
                "payment_mode": "CASH",
                f"return_qty_{invoice_item_id}": "2",
                f"reason_{invoice_item_id}": "Damaged strip",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            ret = self.app_module.Return.query.one()
            return_item = self.app_module.ReturnItem.query.one()
            medicine = self.app_module.Medicine.query.filter_by(name="PARACETAMOL 650", batch="B123").one()
            purchase_item = self.app_module.VendorPurchaseItem.query.one()
            allocation = self.app_module.SalesAllocation.query.one()

            self.assertEqual(ret.invoice_no, "INV-2001")
            self.assertEqual(return_item.qty, 2)
            self.assertEqual(medicine.qty, 2)
            self.assertEqual(purchase_item.remaining_qty, 2)
            self.assertEqual(allocation.returned_qty, 2)

    def test_return_invoice_labels_gst_as_snapshot_and_keeps_refund_total(self):
        with self.app.app_context():
            patient = self._seed_patient()
            invoice = self.app_module.Invoice(
                invoice_no="INV-2002",
                patient_id=patient.id,
                customer=patient.name,
                mobile=patient.mobile,
                subtotal=50.0,
                cgst=1.25,
                sgst=1.25,
                total=50.0,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(invoice)
            self.db.session.flush()
            ret = self.app_module.Return(
                invoice_id=invoice.id,
                invoice_no=invoice.invoice_no,
                customer=invoice.customer,
                mobile=invoice.mobile,
                total_refund=50.0,
                cgst=1.25,
                sgst=1.25,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(ret)
            self.db.session.flush()
            self.db.session.add(
                self.app_module.ReturnItem(
                    return_id=ret.id,
                    invoice_item_id=1,
                    medicine_name="PARACETAMOL 650",
                    qty=2,
                    price=25.0,
                    amount=50.0,
                    discount_percent=0,
                    net_amount=50.0,
                    gst_percent=5.0,
                    reason="Damaged strip",
                )
            )
            self.db.session.commit()
            return_id = ret.id

        self.login()
        response = self.client.get(f"/return-invoice/{return_id}")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("CGST SNAPSHOT (REF)", html)
        self.assertIn("SGST SNAPSHOT (REF)", html)
        self.assertIn("FINAL REFUND TOTAL", html)
        self.assertIn("₹50.0", html)

    def test_appointment_payment_flow_marks_paid_and_blocks_duplicate(self):
        with self.app.app_context():
            patient = self._seed_patient()
            appointment = self.app_module.Appointment(
                appointment_no="APT-1001",
                token_no=1,
                patient_id=patient.id,
                patient_name=patient.name,
                mobile=patient.mobile,
                gender="MALE",
                age=35,
                doctor_name="Dr. Test",
                appointment_date=date.today(),
                appointment_time=time(10, 0),
                payment_mode="ONLINE",
                payment_status="UNPAID",
                consultation_fee=500.0,
                status="BOOKED",
                created_by="admin",
            )
            self.db.session.add(appointment)
            self.db.session.commit()
            appointment_id = appointment.id

        self.login()
        first_response = self.client.post(f"/appointments/{appointment_id}/payment/paid", follow_redirects=False)
        self.assertEqual(first_response.status_code, 302)

        with self.app.app_context():
            appointment = self.app_module.Appointment.query.get(appointment_id)
            self.assertEqual(appointment.payment_status, "PAID")

        second_response = self.client.post(f"/appointments/{appointment_id}/payment/paid", follow_redirects=True)
        self.assertEqual(second_response.status_code, 200)
        self.assertIn(b"already marked as paid", second_response.data)

    def test_appointment_create_handles_legacy_integer_soft_delete_column(self):
        self.login()
        original_helper = self.app_module.appointment_soft_delete_uses_legacy_integer
        self.app_module.appointment_soft_delete_uses_legacy_integer = lambda: True
        try:
            response = self.client.post(
                "/appointments/add",
                data={
                    "patient_name": "Legacy Create",
                    "mobile": "9766655544",
                    "gender": "MALE",
                    "appointment_date": date.today().isoformat(),
                    "appointment_time": "09:45",
                    "payment_mode": "ONLINE",
                    "doctor_discount": "0",
                    "consultation_fee": "600",
                    "symptoms": "",
                    "previous_visit_notes": "",
                    "notes": "",
                },
                follow_redirects=False,
            )
        finally:
            self.app_module.appointment_soft_delete_uses_legacy_integer = original_helper

        self.assertEqual(response.status_code, 302)
        with self.app.app_context():
            appointment = self.app_module.Appointment.query.filter_by(patient_name="Legacy Create").first()
            self.assertIsNotNone(appointment)
            self.assertFalse(bool(appointment.is_deleted))

    def test_profit_report_accuracy(self):
        with self.app.app_context():
            patient = self._seed_patient()
            invoice = self.app_module.Invoice(
                invoice_no="INV-3001",
                patient_id=patient.id,
                customer=patient.name,
                mobile=patient.mobile,
                subtotal=300.0,
                total=300.0,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(invoice)
            self.db.session.flush()
            self.db.session.add(
                self.app_module.InvoiceItem(
                    invoice_id=invoice.id,
                    name="OMEPRAZOLE",
                    qty=3,
                    price=100.0,
                    amount=300.0,
                    batch="OM1",
                    expiry="2027-12-31",
                    net_amount=300.0,
                    cost_price=60.0,
                    cost_amount=180.0,
                )
            )
            ret = self.app_module.Return(
                invoice_id=invoice.id,
                invoice_no=invoice.invoice_no,
                customer=invoice.customer,
                mobile=invoice.mobile,
                total_refund=50.0,
                payment_mode="CASH",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(ret)
            self.db.session.flush()
            self.db.session.add(
                self.app_module.ReturnItem(
                    return_id=ret.id,
                    invoice_item_id=1,
                    medicine_name="OMEPRAZOLE",
                    qty=1,
                    price=50.0,
                    amount=50.0,
                    net_amount=50.0,
                    cost_price=30.0,
                    cost_amount=30.0,
                )
            )
            self.db.session.commit()

            today_iso = date.today().isoformat()
            summary, error = self.app_module.build_profit_report_summary(today_iso, today_iso)
            self.assertIsNone(error)
            self.assertEqual(summary["net_sales"], 250.0)
            self.assertEqual(summary["net_cogs"], 150.0)
            self.assertEqual(summary["gross_profit"], 100.0)

    def test_fifo_stock_deduction_uses_oldest_purchase_cost(self):
        with self.app.app_context():
            self._seed_patient()
            _, _, medicine = self._seed_vendor_purchase_stack(
                total_qty=10,
                purchase_chunks=[
                    (5, 10.0, datetime.utcnow() - timedelta(days=2)),
                    (5, 12.0, datetime.utcnow() - timedelta(days=1)),
                ],
            )

        self.login()
        response = self.client.post(
            "/billing",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "doctor": "Dr. Test",
                "gender": "MALE",
                "payment_mode": "CASH",
                "medicine_name": ["PARACETAMOL 650"],
                "qty": ["6"],
                "batch_override[]": ["B123"],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)

        with self.app.app_context():
            invoice_item = self.app_module.InvoiceItem.query.one()
            purchase_items = self.app_module.VendorPurchaseItem.query.order_by(self.app_module.VendorPurchaseItem.created_at.asc()).all()
            medicine = self.app_module.Medicine.query.filter_by(name="PARACETAMOL 650", batch="B123").one()

            self.assertAlmostEqual(invoice_item.cost_amount, 62.0, places=2)
            self.assertEqual(purchase_items[0].remaining_qty, 0)
            self.assertEqual(purchase_items[1].remaining_qty, 4)
            self.assertEqual(medicine.qty, 4)

    def test_dashboard_profit_cards_hidden_without_permission_and_visible_with_permission(self):
        with self.app.app_context():
            hidden_staff = self.app_module.User(
                username="staff_hidden",
                role="staff",
                access_profile="custom",
                can_invoice_action=True,
            )
            hidden_staff.set_password("Staff@123")
            visible_staff = self.app_module.User(
                username="staff_visible",
                role="staff",
                access_profile="custom",
                can_invoice_action=True,
                can_view_profit_dashboard=True,
            )
            visible_staff.set_password("Staff@123")
            self.db.session.add(hidden_staff)
            self.db.session.add(visible_staff)
            self.db.session.commit()

        self.login_as("staff_hidden", "Staff@123")
        hidden_response = self.client.get("/")
        self.assertEqual(hidden_response.status_code, 200)
        self.assertNotIn(b"Gross Profit Today", hidden_response.data)
        self.assertNotIn(b"Gross Profit This Month", hidden_response.data)

        self.client.get("/logout")
        self.login_as("staff_visible", "Staff@123")
        visible_response = self.client.get("/")
        self.assertEqual(visible_response.status_code, 200)
        self.assertIn(b"Gross Profit Today", visible_response.data)
        self.assertIn(b"Gross Profit This Month", visible_response.data)

    def test_dashboard_expiring_count_excludes_already_expired_stock(self):
        with self.app.app_context():
            today = date.today()
            self.db.session.add_all(
                [
                    self.app_module.Medicine(
                        name="Expired Med",
                        batch="EXP-1",
                        expiry=(today - timedelta(days=1)).isoformat(),
                        mrp=10.0,
                        qty=5,
                        discount_percent=0,
                        is_active=True,
                    ),
                    self.app_module.Medicine(
                        name="Soon Med",
                        batch="SOON-1",
                        expiry=(today + timedelta(days=7)).isoformat(),
                        mrp=10.0,
                        qty=5,
                        discount_percent=0,
                        is_active=True,
                    ),
                    self.app_module.Medicine(
                        name="Later Med",
                        batch="LATE-1",
                        expiry=(today + timedelta(days=45)).isoformat(),
                        mrp=10.0,
                        qty=5,
                        discount_percent=0,
                        is_active=True,
                    ),
                ]
            )
            self.db.session.commit()

        self.login()
        response = self.client.get("/")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertRegex(
            html,
            r'Expiring In 30 Days</div>\s*<div class="app-stat-value">\s*1\s*</div>',
        )

    def test_invalid_login_creates_security_event(self):
        response = self.client.post(
            "/login",
            data={"username": "admin", "password": "WrongPass!"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Invalid credentials", response.data)

        with self.app.app_context():
            event = self.app_module.LoginSecurityEvent.query.order_by(
                self.app_module.LoginSecurityEvent.id.desc()
            ).first()
            self.assertIsNotNone(event)
            self.assertEqual(event.outcome, "FAILED")
            self.assertEqual(event.reason, "Invalid credentials")
            self.assertFalse(event.is_suspicious)

    def test_archived_user_is_soft_deleted_and_login_is_blocked(self):
        with self.app.app_context():
            staff = self.app_module.User(
                username="archivable_staff",
                role="staff",
                access_profile="billing_only",
                can_invoice_action=True,
            )
            staff.set_password("Staff@123")
            self.db.session.add(staff)
            self.db.session.commit()
            staff_id = staff.id

        self.login()
        archive_response = self.client.get(f"/users/delete/{staff_id}", follow_redirects=False)
        self.assertEqual(archive_response.status_code, 302)

        with self.app.app_context():
            archived_user = self.app_module.User.query.get(staff_id)
            self.assertFalse(archived_user.is_active)
            self.assertIsNotNone(archived_user.deleted_at)
            self.assertEqual(self.app_module.active_user_query().filter_by(username="archivable_staff").count(), 0)

        login_response = self.client.post(
            "/login",
            data={"username": "archivable_staff", "password": "Staff@123"},
            follow_redirects=True,
        )
        self.assertEqual(login_response.status_code, 200)
        self.assertIn(b"Invalid credentials", login_response.data)

        with self.app.app_context():
            event = self.app_module.LoginSecurityEvent.query.filter_by(
                username="archivable_staff",
                outcome="FAILED",
            ).order_by(self.app_module.LoginSecurityEvent.id.desc()).first()
            self.assertIsNotNone(event)
            self.assertTrue(event.is_suspicious)
            self.assertEqual(event.reason, "Attempted login to disabled user")

    def test_session_idle_timeout_logs_security_event(self):
        self.login()
        with self.client.session_transaction() as session_data:
            session_data["last_seen_at"] = (datetime.utcnow() - timedelta(hours=3)).isoformat()

        response = self.client.get("/medicines", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn("/login", response.headers.get("Location", ""))

        with self.app.app_context():
            event = self.app_module.LoginSecurityEvent.query.filter_by(
                outcome="SESSION_EXPIRED"
            ).order_by(self.app_module.LoginSecurityEvent.id.desc()).first()
            self.assertIsNotNone(event)
            self.assertEqual(event.reason, "Idle session timeout")

    def test_hold_bill_delete_soft_archives_record(self):
        with self.app.app_context():
            hold_bill = self.app_module.PendingBillStore(
                customer="Saved Patient",
                mobile="9000000000",
                doctor="Dr. Save",
                gender="MALE",
                data_text=self.app_module.serialize_json_text({"items": []}),
            )
            self.db.session.add(hold_bill)
            self.db.session.commit()
            hold_bill_id = hold_bill.id

        self.login()
        response = self.client.get(f"/delete-hold/{hold_bill_id}", follow_redirects=False)
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            hold_bill = self.app_module.PendingBillStore.query.get(hold_bill_id)
            self.assertTrue(hold_bill.is_deleted)
            self.assertIsNotNone(hold_bill.deleted_at)
            self.assertEqual(self.app_module.active_hold_bill_query().filter_by(id=hold_bill_id).count(), 0)

    def test_hold_bill_post_saves_and_redirects(self):
        self.login()
        response = self.client.post(
            "/billing/hold",
            data={
                "customer": "Saved Patient",
                "mobile": "9000000000",
                "doctor": "Dr. Save",
                "gender": "MALE",
                "payment_mode": "CASH",
                "internal_note": "Call before delivery",
                "medicine_name": [""],
                "qty": [""],
                "batch_override[]": [""],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/pending-bills", response.headers.get("Location", ""))

        with self.client.session_transaction() as session_data:
            flashes = session_data.get("_flashes", [])
        self.assertFalse(any("emergency fallback storage" in message.lower() for _, message in flashes))

        with self.app.app_context():
            hold_bill = self.app_module.PendingBillStore.query.one()
            normalized = self.app_module.normalize_hold_bill_data(hold_bill)
            self.assertEqual(hold_bill.customer, "Saved Patient")
            self.assertEqual(normalized["header"]["doctor"], "Dr. Save")
            self.assertEqual(normalized["header"]["payment_mode"], "CASH")
            self.assertEqual(normalized["header"]["internal_note"], "Call before delivery")

    def test_reports_invoice_report_shows_internal_note(self):
        with self.app.app_context():
            patient = self._seed_patient()
            invoice = self.app_module.Invoice(
                invoice_no="INV-REPORT-1",
                patient_id=patient.id,
                customer=patient.name,
                mobile=patient.mobile,
                subtotal=50.0,
                total=50.0,
                payment_mode="CASH",
                internal_note="Doctor sample adjustment",
                created_by="admin",
                created_at=datetime.utcnow(),
            )
            self.db.session.add(invoice)
            self.db.session.commit()

        self.login()
        response = self.client.post(
            "/reports",
            data={"report_type": "daily"},
            follow_redirects=True,
        )
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Internal Note", response.data)
        self.assertIn(b"Doctor sample adjustment", response.data)

    def test_reports_page_defaults_to_cutover_window_but_keeps_older_history_accessible(self):
        with self.app.app_context():
            self._seed_invoice(
                customer="Legacy Report Patient",
                mobile="9000000021",
                total=120.0,
                created_at=datetime(2026, 6, 30, 12, 0, 0),
            )
            _patient, current_invoice = self._seed_invoice(
                customer="Current Report Patient",
                mobile="9000000022",
                total=180.0,
                created_at=datetime(2026, 7, 2, 12, 0, 0),
            )
            current_invoice_no = current_invoice.invoice_no

        self.login()
        response = self.client.get("/reports")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('value="2026-07-01"', html)
        self.assertIn(current_invoice_no, html)
        self.assertNotIn("Legacy Report Patient", html)

        legacy_response = self.client.post(
            "/reports",
            data={
                "report_type": "custom",
                "from_date": "2026-06-01",
                "to_date": "2026-06-30",
            },
            follow_redirects=True,
        )
        self.assertEqual(legacy_response.status_code, 200)
        self.assertIn(b"Legacy Report Patient", legacy_response.data)

    def test_invoice_list_defaults_to_cutover_window_but_allows_older_history_filters(self):
        with self.app.app_context():
            self._seed_invoice(
                customer="Legacy Invoice Patient",
                mobile="9000000031",
                total=95.0,
                created_at=datetime(2026, 6, 30, 9, 0, 0),
            )
            _patient, current_invoice = self._seed_invoice(
                customer="Current Invoice Patient",
                mobile="9000000032",
                total=205.0,
                created_at=datetime(2026, 7, 2, 9, 0, 0),
            )
            current_invoice_no = current_invoice.invoice_no

        self.login()
        response = self.client.get("/invoices")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn('value="2026-07-01"', html)
        self.assertIn(current_invoice_no, html)
        self.assertNotIn("Legacy Invoice Patient", html)

        legacy_response = self.client.get("/invoices?from=2026-06-01&to=2026-06-30")
        self.assertEqual(legacy_response.status_code, 200)
        self.assertIn(b"Legacy Invoice Patient", legacy_response.data)

    def test_full_reports_export_medicines_sheet_includes_purchase_rate(self):
        with self.app.app_context():
            self._seed_vendor_purchase_stack(total_qty=8, purchase_chunks=[(8, 13.75, datetime.utcnow())])

        self.login()
        response = self.client.get("/reports/export?scope=all")
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.data))
        sheet = workbook["Medicines"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Purchase Rate", headers)
        purchase_rate_index = headers.index("Purchase Rate") + 1
        self.assertEqual(sheet.cell(row=2, column=purchase_rate_index).value, 13.75)

    def test_medicines_export_includes_purchase_rate_column(self):
        with self.app.app_context():
            self._seed_vendor_purchase_stack(total_qty=8, purchase_chunks=[(8, 17.25, datetime.utcnow())])

        self.login()
        response = self.client.get("/medicines/export")
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(BytesIO(response.data))
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Purchase Rate", headers)
        purchase_rate_index = headers.index("Purchase Rate") + 1
        self.assertEqual(sheet.cell(row=2, column=purchase_rate_index).value, 17.25)

    def test_admin_invoice_edit_can_move_cash_invoice_to_online_mode(self):
        with self.app.app_context():
            _patient, invoice = self._seed_invoice(total=50.0, payment_mode="CASH", cash_amount=50.0, online_amount=0.0)
            invoice_id = invoice.id

        self.login()
        response = self.client.post(
            f"/invoice/edit/{invoice_id}",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "payment_mode": "UPI",
                "internal_note": "Corrected by admin",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(f"/invoice/{invoice_id}", response.headers.get("Location", ""))

        with self.app.app_context():
            updated_invoice = self.db.session.get(self.app_module.Invoice, invoice_id)
            self.assertEqual(updated_invoice.payment_mode, "UPI")
            self.assertEqual(float(updated_invoice.cash_amount), 0.0)
            self.assertEqual(float(updated_invoice.online_amount), 50.0)
            self.assertFalse(updated_invoice.is_split_payment)
            self.assertEqual(updated_invoice.internal_note, "Corrected by admin")

    def test_admin_invoice_edit_preserves_split_breakdown_when_online_mode_changes(self):
        with self.app.app_context():
            _patient, invoice = self._seed_invoice(
                total=50.0,
                payment_mode="UPI",
                cash_amount=20.0,
                online_amount=30.0,
                is_split_payment=True,
            )
            invoice_id = invoice.id

        self.login()
        response = self.client.post(
            f"/invoice/edit/{invoice_id}",
            data={
                "customer": "Ravi Kumar",
                "mobile": "9876543210",
                "payment_mode": "CARD",
                "internal_note": "Mode corrected",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn(f"/invoice/{invoice_id}", response.headers.get("Location", ""))

        with self.app.app_context():
            updated_invoice = self.db.session.get(self.app_module.Invoice, invoice_id)
            self.assertEqual(updated_invoice.payment_mode, "CARD")
            self.assertEqual(float(updated_invoice.cash_amount), 20.0)
            self.assertEqual(float(updated_invoice.online_amount), 30.0)
            self.assertTrue(updated_invoice.is_split_payment)
            self.assertEqual(updated_invoice.internal_note, "Mode corrected")

    def test_non_admin_users_cannot_edit_invoices_and_do_not_see_edit_action(self):
        with self.app.app_context():
            _patient, invoice = self._seed_invoice(total=50.0, payment_mode="CASH", cash_amount=50.0, online_amount=0.0)
            invoice_id = invoice.id
            staff_user = self.app_module.User(
                username="billing_staff",
                role="staff",
                access_profile="custom",
                can_invoice_action=True,
                can_edit_invoice=True,
                is_active=True,
            )
            staff_user.set_password("Password@123")
            self.db.session.add(staff_user)
            self.db.session.commit()

        self.login_as("billing_staff", "Password@123")
        list_response = self.client.get("/invoices")
        self.assertEqual(list_response.status_code, 200)
        self.assertNotIn(f'/invoice/edit/{invoice_id}'.encode(), list_response.data)

        edit_response = self.client.get(f"/invoice/edit/{invoice_id}", follow_redirects=False)
        self.assertEqual(edit_response.status_code, 302)
        self.assertIn("/invoices", edit_response.headers.get("Location", ""))

    def test_normalize_hold_bill_data_accepts_string_payload(self):
        payload = {
            "header": {
                "customer": "Legacy Patient",
                "mobile": "9000000000",
                "doctor": "Dr. Legacy",
                "gender": "FEMALE",
                "sale_type": "sale",
                "payment_mode": "UPI",
            },
            "items": [
                {
                    "name": "DOLO 650",
                    "qty": 2,
                    "mrp": 30,
                    "discount_percent": 0,
                    "net_amount": 60,
                }
            ],
            "totals": {
                "subtotal": 60,
                "discount": 0,
                "cgst": 1.5,
                "sgst": 1.5,
                "net_total": 60,
                "rounded_amount": 60,
            },
        }

        with self.app.app_context():
            hold_bill = self.app_module.PendingBillStore(
                customer="Legacy Patient",
                mobile="9000000000",
                doctor="Dr. Legacy",
                gender="FEMALE",
                data_text=self.app_module.serialize_json_text(payload),
            )
            self.db.session.add(hold_bill)
            self.db.session.commit()
            hold_bill_id = hold_bill.id

            saved_hold_bill = self.app_module.PendingBillStore.query.get(hold_bill_id)
            normalized = self.app_module.normalize_hold_bill_data(saved_hold_bill)

        self.assertEqual(normalized["header"]["customer"], "Legacy Patient")
        self.assertEqual(normalized["header"]["payment_mode"], "UPI")
        self.assertEqual(len(normalized["items"]), 1)
        self.assertEqual(normalized["items"][0]["name"], "DOLO 650")

    def test_hold_bill_post_sanitizes_nan_numeric_values(self):
        self.login()
        response = self.client.post(
            "/billing/hold",
            data={
                "customer": "Nan Patient",
                "mobile": "9000000000",
                "doctor": "Dr. Nan",
                "gender": "MALE",
                "payment_mode": "CASH",
                "medicine_name": ["DOLO 650"],
                "qty": ["1"],
                "batch_override[]": [""],
                "line_qoh[]": ["NaN"],
                "line_mrp[]": ["NaN"],
                "line_discount_percent[]": ["NaN"],
                "line_net[]": ["NaN"],
                "subtotal": "NaN",
                "discount": "NaN",
                "cgst": "NaN",
                "sgst": "NaN",
                "net_total": "NaN",
                "rounded_amount": "NaN",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/pending-bills", response.headers.get("Location", ""))

        with self.app.app_context():
            hold_bill = self.app_module.PendingBillStore.query.order_by(self.app_module.PendingBillStore.id.desc()).first()
            normalized = self.app_module.normalize_hold_bill_data(hold_bill)

        self.assertEqual(normalized["totals"]["subtotal"], 0.0)
        self.assertEqual(normalized["totals"]["net_total"], 0.0)
        self.assertEqual(normalized["items"][0]["mrp"], 0.0)
        self.assertEqual(normalized["items"][0]["net_amount"], 0.0)

    def test_hold_bill_post_recovers_when_table_is_missing(self):
        with self.app.app_context():
            self.db.session.execute(self.app_module.text('DROP TABLE IF EXISTS "pending_bill_store"'))
            self.db.session.commit()

        self.login()
        response = self.client.post(
            "/billing/hold",
            data={
                "customer": "Recovered Patient",
                "mobile": "9111111111",
                "doctor": "Dr. Recovery",
                "gender": "Male",
                "payment_mode": "CASH",
                "medicine_name": [""],
                "qty": [""],
                "batch_override[]": [""],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn("/pending-bills", response.headers.get("Location", ""))

        with self.app.app_context():
            rows = self.db.session.execute(
                self.app_module.text('SELECT customer FROM "pending_bill_store" ORDER BY id DESC LIMIT 1')
            ).fetchall()
        self.assertEqual(rows[0][0], "Recovered Patient")

    def test_pending_bills_route_migrates_file_fallback_records_into_store(self):
        with self.app.app_context():
            self.app_module.save_file_hold_bill(
                customer="File Patient",
                mobile="9222222222",
                doctor="Dr. File",
                gender="Male",
                payload={"items": [], "header": {"doctor": "Dr. File"}},
            )
        self.login()
        response = self.client.get("/pending-bills", follow_redirects=False)
        self.assertEqual(response.status_code, 200)

        with self.app.app_context():
            self.assertEqual(self.app_module.list_file_hold_bills(), [])
            hold_bill = self.app_module.PendingBillStore.query.filter_by(customer="File Patient").one()
            restored = self.app_module.normalize_hold_bill_data(hold_bill)
        self.assertEqual(hold_bill.customer, "File Patient")
        self.assertEqual(restored["header"]["doctor"], "Dr. File")

    def test_appointment_delete_soft_archives_record(self):
        with self.app.app_context():
            patient = self._seed_patient(name="Delete Appt", mobile="9888877777")
            appointment = self.app_module.Appointment(
                appointment_no="APT-DEL-1",
                token_no=3,
                patient_id=patient.id,
                patient_name=patient.name,
                mobile=patient.mobile,
                gender="MALE",
                age=34,
                doctor_name="Dr. Delete",
                appointment_date=date.today(),
                appointment_time=time(12, 0),
                consultation_fee=400.0,
                payment_mode="CASH",
                payment_status="UNPAID",
                status="BOOKED",
                created_by="admin",
            )
            self.db.session.add(appointment)
            self.db.session.commit()
            appointment_id = appointment.id

        self.login()
        response = self.client.post(f"/appointments/delete/{appointment_id}", follow_redirects=False)
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            appointment = self.app_module.Appointment.query.get(appointment_id)
            self.assertTrue(appointment.is_deleted)
            self.assertIsNotNone(appointment.deleted_at)
            self.assertEqual(self.app_module.active_appointment_query().filter_by(id=appointment_id).count(), 0)

    def test_appointment_delete_async_returns_success_payload(self):
        with self.app.app_context():
            patient = self._seed_patient(name="Delete Async", mobile="9777766666")
            appointment = self.app_module.Appointment(
                appointment_no="APT-DEL-ASYNC",
                token_no=4,
                patient_id=patient.id,
                patient_name=patient.name,
                mobile=patient.mobile,
                gender="MALE",
                age=31,
                doctor_name="Dr. Delete",
                appointment_date=date.today(),
                appointment_time=time(1, 0),
                consultation_fee=500.0,
                payment_mode="ONLINE",
                payment_status="UNPAID",
                status="BOOKED",
                created_by="admin",
            )
            self.db.session.add(appointment)
            self.db.session.commit()
            appointment_id = appointment.id

        self.login()
        response = self.client.post(
            f"/appointments/delete/{appointment_id}",
            headers={"X-Requested-With": "XMLHttpRequest"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])

        with self.app.app_context():
            appointment = self.app_module.Appointment.query.get(appointment_id)
            self.assertTrue(appointment.is_deleted)
            self.assertEqual(self.app_module.active_appointment_query().filter_by(id=appointment_id).count(), 0)

    def test_user_edit_page_renders_and_updates_permissions(self):
        with self.app.app_context():
            staff_user = self.app_module.User(
                username="counter1",
                role="staff",
                access_profile="custom",
                can_view_reports=False,
                can_manage_purchases=False,
                can_view_profit_dashboard=False,
                is_active=True,
            )
            staff_user.set_password("Password@123")
            self.db.session.add(staff_user)
            self.db.session.commit()
            staff_user_id = staff_user.id

        self.login()
        get_response = self.client.get(f"/users/edit/{staff_user_id}")
        self.assertEqual(get_response.status_code, 200)
        self.assertIn(b"Dashboard Profit Cards", get_response.data)
        self.assertIn(b"Access Profile", get_response.data)

        response = self.client.post(
            f"/users/edit/{staff_user_id}",
            data={
                "username": "counter1",
                "role": "staff",
                "access_profile": "custom",
                "can_view_reports": "on",
                "can_manage_purchases": "on",
                "can_view_profit_dashboard": "on",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.app.app_context():
            updated_user = self.db.session.get(self.app_module.User, staff_user_id)
            self.assertTrue(updated_user.can_view_reports)
            self.assertTrue(updated_user.can_manage_purchases)
            self.assertTrue(updated_user.can_view_profit_dashboard)
            self.assertEqual(updated_user.access_profile, "custom")

    def test_user_edit_and_create_survive_legacy_integer_permission_storage(self):
        with self.app.app_context():
            staff_user = self.app_module.User(
                username="legacy_staff",
                role="staff",
                access_profile="custom",
                can_manage_purchases=False,
                can_view_reports=False,
                is_active=True,
            )
            staff_user.set_password("Password@123")
            self.db.session.add(staff_user)
            self.db.session.commit()
            staff_user_id = staff_user.id

        self.login()
        original_helper = self.app_module.user_boolean_storage_mode_map
        legacy_map = {"is_active": "integer"}
        legacy_map.update({field_name: "integer" for field_name in self.app_module.USER_PERMISSION_FIELDS})
        self.app_module.user_boolean_storage_mode_map = lambda: dict(legacy_map)
        try:
            edit_response = self.client.post(
                f"/users/edit/{staff_user_id}",
                data={
                    "username": "legacy_staff",
                    "role": "staff",
                    "access_profile": "custom",
                    "can_manage_purchases": "on",
                    "can_view_reports": "on",
                },
                follow_redirects=False,
            )
            self.assertEqual(edit_response.status_code, 302)

            add_response = self.client.post(
                "/users/add",
                data={
                    "username": "legacy_new",
                    "password": "Password@123",
                    "role": "staff",
                    "access_profile": "custom",
                    "can_view_reports": "on",
                    "can_manage_purchases": "on",
                },
                follow_redirects=False,
            )
            self.assertEqual(add_response.status_code, 302)
        finally:
            self.app_module.user_boolean_storage_mode_map = original_helper

        with self.app.app_context():
            updated_user = self.db.session.get(self.app_module.User, staff_user_id)
            created_user = self.app_module.active_user_query().filter_by(username="legacy_new").first()
            self.assertTrue(updated_user.can_manage_purchases)
            self.assertTrue(updated_user.can_view_reports)
            self.assertIsNotNone(created_user)
            self.assertTrue(created_user.can_manage_purchases)
            self.assertTrue(created_user.can_view_reports)

    def test_user_delete_survives_legacy_integer_is_active_storage(self):
        with self.app.app_context():
            staff_user = self.app_module.User(
                username="archivable_staff",
                role="staff",
                access_profile="custom",
                is_active=True,
            )
            staff_user.set_password("Password@123")
            self.db.session.add(staff_user)
            self.db.session.commit()
            staff_user_id = staff_user.id

        self.login()
        original_helper = self.app_module.user_boolean_storage_mode_map
        self.app_module.user_boolean_storage_mode_map = lambda: {"is_active": "integer"}
        try:
            response = self.client.get(f"/users/delete/{staff_user_id}", follow_redirects=False)
            self.assertEqual(response.status_code, 302)
        finally:
            self.app_module.user_boolean_storage_mode_map = original_helper

        with self.app.app_context():
            archived_user = self.db.session.get(self.app_module.User, staff_user_id)
            self.assertIsNotNone(archived_user.deleted_at)
            self.assertEqual(self.app_module.active_user_query().filter_by(username="archivable_staff").count(), 0)

    def test_backup_snapshot_and_restore_drill(self):
        with self.app.app_context():
            upload_dirs = self.app.config["INFRA_UPLOAD_DIRECTORIES"]
            os.makedirs(upload_dirs["vendor_uploads"], exist_ok=True)
            with open(os.path.join(upload_dirs["vendor_uploads"], "sample.txt"), "w", encoding="utf-8") as handle:
                handle.write("backup-check")

            snapshot = self.app_module.build_backup_snapshot(
                self.app,
                upload_dirs=upload_dirs,
                keep_count=5,
                include_uploads=True,
            )
            self.assertTrue(os.path.exists(snapshot["manifest_path"]))
            self.assertTrue(os.path.exists(snapshot["restore_plan_path"]))

            drill = self.app_module.restore_backup_snapshot(
                self.app,
                snapshot_name=snapshot["snapshot_name"],
                include_uploads=True,
                dry_run=True,
            )
            self.assertTrue(drill["dry_run"])
            self.assertEqual(drill["snapshot_name"], snapshot["snapshot_name"])


if __name__ == "__main__":
    unittest.main()
